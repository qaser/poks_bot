import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from config.mongo_config import gpa, reqs


def create_report_excel():
    queryset = reqs.find({})

    # Создаем рабочую книгу и лист
    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки на ТО"

    # Заголовки столбцов
    headers = [
        "№ заявки",
        "Тип заявки",
        "Статус",
        "Дата создания заявки",
        "Планируемая дата пуска",
        "Год",
        "Месяц",
        "ЛПУМГ",
        "КС",
        "Газопровод",
        "КЦ",
        "№ ГПА",
        "Наименование ГПА",
        "Тип ГПА",
        "Группа ГПА",
        "Тип ЦБН",
        "Тип двигателя",
        "МРР",
        "Приоритет",
        "Критерий приоритета",
        "Отказ при пуске",
        "Причина отказа при пуске",
        "Причина отклонения заявки"
    ]

    # Записываем заголовки с увеличенной высотой
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # Устанавливаем высоту строки для заголовков
    ws.row_dimensions[1].height = 40

    # Стили для ячеек
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Заполняем данные
    row_num = 2
    for req in queryset:
        # Получаем данные ГПА
        gpa_instance = gpa.find_one({'_id': req['gpa_id']})

        # Форматируем данные
        req_num = req.get('req_num', '')

        # Преобразуем тип заявки в читабельный формат
        req_type = req.get('req_type', '')
        if req_type == 'with_approval':
            req_type_display = 'Плановая'
        elif req_type == 'without_approval':
            req_type_display = 'Без согласования'
        else:
            req_type_display = req_type

        # Преобразуем статус в читабельный формат и определяем цвет
        status = req.get('status', '')
        status_color = None
        if status == 'inwork':
            status_display = 'В работе'
            status_color = "FFEB9C"  # Мягкий желтый (светло-желтый)
        elif status == 'approved':
            status_display = 'Согласовано'
            status_color = "C6EFCE"  # Мягкий зеленый (светло-зеленый)
        elif status == 'rejected':
            status_display = 'Отклонено'
            status_color = "FFC7CE"  # Мягкий красный (светло-розовый)
        else:
            status_display = status

        ks = req.get('ks', '')
        num_gpa = req.get('num_gpa', '')

        # Данные ГПА
        num_shop = gpa_instance.get('num_shop', '') if gpa_instance else ''
        name_gpa = gpa_instance.get('name_gpa', '') if gpa_instance else ''
        type_gpa = gpa_instance.get('type_gpa', '') if gpa_instance else ''
        group_gpa = gpa_instance.get('group_gpa', '') if gpa_instance else ''
        cbn_type = gpa_instance.get('cbn_type', '') if gpa_instance else ''
        engine_type = gpa_instance.get('engine_type', '') if gpa_instance else ''
        lpu = gpa_instance.get('lpu', '') if gpa_instance else ''
        pipeline = gpa_instance.get('pipeline', '') if gpa_instance else ''

        # Даты
        create_date = req.get('datetime')
        if create_date and isinstance(create_date, datetime):
            create_date_str = create_date.strftime('%d.%m.%Y %H:%M')
        else:
            create_date_str = ''

        plan_date = req.get('request_datetime')
        if plan_date and isinstance(plan_date, datetime):
            plan_date_str = plan_date.strftime('%d.%m.%Y %H:%M')
            year = plan_date.year
            month = plan_date.month
        else:
            plan_date_str = ''
            year = ''
            month = ''

        # Статусы и флаги
        resource = req.get('resource', '')
        priority = req.get('priority', '')
        priority_criteria = req.get('priority_criteria', '')
        is_fail = 'Да' if req.get('is_fail') else 'Нет'
        fail_reason = req.get('fail_reason', '')
        reject_reason = req.get('reject_reason', '')

        # Записываем данные в строку
        data_row = [
            req_num,
            req_type_display,
            status_display,  # Статус для отображения
            create_date_str,
            plan_date_str,
            year,
            month,
            lpu,
            ks,
            pipeline,
            num_shop,
            num_gpa,
            name_gpa,
            type_gpa,
            group_gpa,
            cbn_type,
            engine_type,
            resource,
            priority,
            priority_criteria,
            is_fail,
            fail_reason,
            reject_reason
        ]

        for col_num, value in enumerate(data_row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border

            # Применяем цвет для ячейки статуса
            if col_num == 3 and status_color:
                cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")

            if isinstance(value, (int, float)) or (isinstance(value, str) and value.isdigit()):
                cell.alignment = center_alignment
            else:
                cell.alignment = center_alignment

        row_num += 1

    # УВЕЛИЧИВАЕМ ВЫСОТУ ВСЕХ СТРОК С ДАННЫМИ
    for row in range(2, row_num):  # От 2 строки до последней заполненной
        ws.row_dimensions[row].height = 24  # Высота 20 пунктов

    # Настраиваем ширину столбцов
    column_widths = [10, 20, 20, 20, 20, 8, 8, 20, 20, 20, 8, 8, 20, 20, 20, 20, 20, 12, 16, 16, 20, 20, 20]

    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Замораживаем первую строку с заголовками
    ws.freeze_panes = "A2"

    # Сохраняем файл
    filename = f"static/docs_email/Заявки на пуск ГПА на {datetime.now().strftime('%d.%m.%Y')}.xlsx"
    filepath = os.path.join(os.getcwd(), filename)

    wb.save(filepath)

    return filepath
