import datetime as dt

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import utils.constants as const
from config.mongo_config import gpa, operating_time


def create_report_excel(queryset, date):
    workbook = Workbook()
    worksheet = workbook.active
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrapped_alignment = Alignment(vertical='top', wrap_text=True)
    title = f'Информация по наработке ГПА за {const.MONTHS_NAMES[str(date.month)]} {date.year} г.'
    worksheet.title = f'{date.month} {date.year}'
    columns = [
        ('№ п/п', 5),
        ('Компрессорная станция', 40),
        ('№ ГПА', 20),
        ('Наработка, ч.', 20),
        ('Суммарная наработка по КС, ч.', 20)
    ]
    worksheet.merge_cells(start_column=1, start_row=1, end_column=len(columns), end_row=1)
    title_cell = worksheet.cell(row=1, column=1)
    title_cell.value = title
    title_cell.alignment = centered_alignment
    title_cell.font = header_font
    worksheet.row_dimensions[1].height = 30
    row_num = 2  # начинаем заполнять со второй строки, т.к. первая сторока - заголовок
    merge_row = row_num + 1
    sum_time = 0
    for col_num, (column_title, column_width) in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.border = thin_border
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = column_width
        worksheet.row_dimensions[row_num].height = 30
    for num, ks_data in enumerate(queryset):
        gpa_ids = ks_data['gpa_ids']
        ks = ks_data['ks']
        sum_ks_time = 0
        for gpa_id in gpa_ids:
            row_num += 1
            work_time = operating_time.find_one(
                {'gpa_id': gpa_id, 'month': date.month, 'year': date.year}
            )['work_time']
            num_gpa = gpa.find_one({'_id': gpa_id})['num_gpa']
            row = [
                num + 1,
                ks,
                num_gpa,
                work_time,
                # 0,
            ]
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                worksheet.row_dimensions[row_num].height = 20
                cell.value = cell_value
                cell.alignment = centered_alignment
                cell.border = thin_border
            sum_time += work_time
            sum_ks_time += work_time
        time_cell = worksheet.cell(row=merge_row, column=5)
        time_cell.value = sum_ks_time
        time_cell.alignment = centered_alignment
        time_cell.border = thin_border
        worksheet.merge_cells(start_column=1, start_row=merge_row, end_column=1, end_row=merge_row+len(gpa_ids)-1)
        worksheet.merge_cells(start_column=2, start_row=merge_row, end_column=2, end_row=merge_row+len(gpa_ids)-1)
        worksheet.merge_cells(start_column=5, start_row=merge_row, end_column=5, end_row=merge_row+len(gpa_ids)-1)
        merge_row += len(gpa_ids)
        sum_time_cell = worksheet.cell(row=merge_row, column=5)
        sum_time_cell.value = sum_time
        sum_time_cell.alignment = centered_alignment
        sum_time_cell.border = thin_border
    path = f'static/docs_email/Отчет по наработке ГПА.xlsx'
    workbook.save(path)
    return path
