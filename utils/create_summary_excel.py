import datetime as dt

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import utils.constants as const
from config.mongo_config import petitions


def create_summary_excel(period):
    workbook = Workbook()
    worksheet = workbook.active
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrapped_alignment = Alignment(vertical='top', wrap_text=True)
    worksheet.title = 'Отчет по проблемным вопросам'
    columns = [
        ('№ п/п', 5),
        ('КС', 40),
        ('Направление', 30),
        ('Дата', 20),
        ('Содержание вопроса', 125),
        ('Статус', 20),
    ]
    worksheet.merge_cells(start_column=1, start_row=1, end_column=len(columns), end_row=1)
    title_cell = worksheet.cell(row=1, column=1)
    if period == 'week':
        title = 'Отчет по проблемным вопросам филиалов по направлению ПОпоЭКС за неделю'
    elif period == 'month':
        title = 'Сводный отчет по проблемным вопросам филиалов по направлению ПОпоЭКС'
    title_cell.value = title
    title_cell.alignment = centered_alignment
    title_cell.font = header_font
    worksheet.row_dimensions[1].height = 30
    row_num = 2  # начинаем заполнять со второй строки, т.к. первая сторока - заголовок
    for col_num, (column_title, column_width) in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.border = thin_border
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = column_width
        worksheet.row_dimensions[row_num].height = 30
    for ks in const.KS:
        # нерешенные вопросы и решенные в течение этой недели
        if period == 'week':
            now = dt.datetime.now()
            previous_week = now - dt.timedelta(days=7)
            title = 'Отчет по проблемным вопросам филиалов по направлению ПОпоЭКС за неделю'
            pipeline = {'$or': [
                {'ks': ks, 'status': {'$in': ['inwork', 'create']}},
                {'ks': ks, 'status': 'finish', 'date': {"$lte": now, "$gte": previous_week}}
            ]}
        elif period == 'month':
            title = 'Отчет по проблемным вопросам филиалов по направлению ПОпоЭКС'
            pipeline = {'ks': ks, 'status': {'$in': ['inwork', 'create', 'rework', 'finish', 'delete']}}
        queryset = list(petitions.find(pipeline).sort([('ks', 1), ('status', 1), ('directions', 1)]))
        num_petitions = len(queryset)
        if num_petitions == 0:
            continue
        for num, pet in enumerate(queryset):
            row_num += 1
            status = pet.get('status')
            color = const.PETITION_COLOR[status]
            row = [
                str(num + 1),
                ks,
                const.DIRECTIONS_CODES[pet.get('direction')],
                pet.get('date').strftime('%d.%m.%Y %H:%M'),
                pet.get('text'),
                const.PETITION_STATUS[status][0]
            ]
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                worksheet.row_dimensions[row_num].height = 50
                cell.value = cell_value
                cell.alignment = wrapped_alignment
                if col_num == len(row):
                    cell_color = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    cell.fill = cell_color
                cell.border = thin_border
    path = f'static/docs_email/Сводный перечень вопросов.xlsx'
    workbook.save(path)
